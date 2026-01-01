import streamlit as st
import yfinance as yf
import pandas as pd
import numpy as np
import ta
from datetime import datetime, timedelta
import pytz
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import time
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
import warnings
import json
from bs4 import BeautifulSoup
import re
from io import BytesIO
import openpyxl
from sklearn.cluster import KMeans
from scipy.signal import argrelextrema
warnings.filterwarnings('ignore')

# Set page config
st.set_page_config(
    page_title="NSE F&O PCS Professional Scanner",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="auto"  # Collapsible on mobile, expanded on desktop
)

# ENHANCED PROFESSIONAL UI SYSTEM - Tailwind-Inspired CSS
# BLOOMBERG TERMINAL PROFESSIONAL UI SYSTEM
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Roboto+Mono:wght@400;500;600;700&display=swap');
    
    :root {
        /* === BLOOMBERG TERMINAL DESIGN SYSTEM === */
        /* Professional Financial Terminal Theme */
        
        /* Bloomberg Signature Colors */
        --bloomberg-orange: #ff6b00;
        --bloomberg-orange-hover: #ff8533;
        --bloomberg-orange-dark: #cc5500;
        --bloomberg-green: #00d084;
        --bloomberg-red: #ff3b69;
        --bloomberg-blue: #00a3ff;
        
        /* Dark Theme Palette */
        --bg-primary: #0a0a0a;
        --bg-secondary: #151515;
        --bg-elevated: #1a1a1a;
        --bg-hover: #202020;
        
        /* Border System */
        --border-primary: #2a2a2a;
        --border-secondary: #3333;
        --border-accent: var(--bloomberg-orange);
        
        /* Text System */
        --text-primary: #ffff;
        --text-secondary: #9999;
        --text-tertiary: #6666;
        --text-muted: #4444;
        
        /* Semantic Colors */
        --success: var(--bloomberg-green);
        --error: var(--bloomberg-red);
        --warning: #ffaa00;
        --info: var(--bloomberg-blue);
        
        /* Shadows */
        --shadow-sm: 0 1px 2px 0 rgba(0, 0, 0, 0.3);
        --shadow-md: 0 2px 4px 0 rgba(0, 0, 0, 0.4);
        --shadow-lg: 0 4px 8px 0 rgba(0, 0, 0, 0.5);
        
        /* Spacing (keeping existing) */
        --spacing-1: 0.25rem;
        --spacing-2: 0.5rem;
        --spacing-3: 0.75rem;
        --spacing-4: 1rem;
        --spacing-5: 1.25rem;
        --spacing-6: 1.5rem;
        --spacing-8: 2rem;
        
        /* Typography */
        --font-mono: 'Roboto Mono', 'Courier New', monospace;
        --font-size-xs: 0.7rem;
        --font-size-sm: 0.8rem;
        --font-size-base: 0.9rem;
        --font-size-lg: 1rem;
        --font-size-xl: 1.2rem;
        --font-size-2xl: 1.5rem;
    }
    
    /* === BLOOMBERG GLOBAL STYLES === */
    * {
        font-family: var(--font-mono);
        -webkit-font-smoothing: antialiased;
        -moz-osx-font-smoothing: grayscale;
    }
    
    /* Main App Container - Bloomberg Dark */
    .main {
        background: var(--bg-primary) !important;
        color: var(--text-primary) !important;
        padding: var(--spacing-6) var(--spacing-8);
    }
    
    .stApp {
        background: var(--bg-primary) !important;
    }
    
    /* === HEADERS - Bloomberg Style === */
    h1, h2, h3, h4, h5, h6 {
        color: var(--text-primary) !important;
        font-weight: 600;
        letter-spacing: 0.5px;
        font-family: var(--font-mono) !important;
    }
    
    h1 {
        font-size: var(--font-size-2xl);
        margin-bottom: var(--spacing-4);
        color: var(--bloomberg-orange) !important;
        text-transform: uppercase;
        border-bottom: 2px solid var(--border-primary);
        padding-bottom: var(--spacing-2);
    }
    
    h2 {
        font-size: var(--font-size-xl);
        margin-bottom: var(--spacing-3);
        color: var(--bloomberg-orange) !important;
    }
    
    h3 {
        font-size: var(--font-size-lg);
        margin-bottom: var(--spacing-2);
        color: var(--text-secondary) !important;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
    
    /* === SIDEBAR - Bloomberg Terminal Style === */
    [data-testid="stSidebar"] {
        min-width: 342px !important;
        max-width: 342px !important;
        background: var(--bg-secondary) !important;
        border-right: 3px solid var(--bloomberg-orange) !important;
        box-shadow: var(--shadow-lg) !important;
    }
    
    [data-testid="stSidebar"] > div:first-child {
        background: transparent !important;
    }
    
    /* Remove card backgrounds from sidebar elements */
    [data-testid="stSidebar"] .element-container,
    [data-testid="stSidebar"] .stMarkdown,
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
    }
    
    /* Sidebar Text */
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] div {
        color: var(--text-primary) !important;
        font-size: var(--font-size-sm) !important;
    }
    
    /* Sidebar Headers */
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 {
        color: var(--bloomberg-orange) !important;
        font-size: var(--font-size-base) !important;
    }
    
    /* Sidebar Checkboxes */
    [data-testid="stSidebar"] [data-testid="stCheckbox"] label {
        color: var(--text-secondary) !important;
        font-size: var(--font-size-xs) !important;
    }
    
    /* === SIDEBAR COLLAPSE BUTTON - Always Visible & Styled === */
    [data-testid="collapsedControl"] {
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        background: rgba(255, 107, 0, 0.15) !important;
        border: 2px solid var(--bloomberg-orange) !important;
        border-radius: 0 12px 12px 0 !important;
        padding: 12px 8px !important;
        cursor: pointer !important;
        transition: all 0.3s ease !important;
        box-shadow: 2px 2px 8px rgba(0, 0, 0, 0.3) !important;
    }
    
    [data-testid="collapsedControl"]:hover {
        background: var(--bloomberg-orange) !important;
        border-color: var(--bloomberg-orange-hover) !important;
        box-shadow: 2px 2px 12px rgba(255, 107, 0, 0.5) !important;
    }
    
    [data-testid="collapsedControl"] svg {
        fill: var(--bloomberg-orange) !important;
        stroke: var(--bloomberg-orange) !important;
    }
    
    [data-testid="collapsedControl"]:hover svg {
        fill: white !important;
        stroke: white !important;
    }
    
    /* Sidebar collapsed state */
    [data-testid="stSidebar"][aria-expanded="false"] {
        display: none !important;
    }
    
    /* Main content adjustment when sidebar collapsed */
    .main[data-sidebar-collapsed="true"] {
        margin-left: 0 !important;
        max-width: 100% !important;
    }
    
    /* === INPUT FIELDS - Bloomberg Style === */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div > select,
    .stSlider > div > div > div > div {
        background: var(--bg-secondary) !important;
        border: 1px solid var(--border-primary) !important;
        border-radius: 2px !important;
        color: var(--text-primary) !important;
        font-size: var(--font-size-sm) !important;
        font-family: var(--font-mono) !important;
    }
    
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus,
    .stSelectbox > div > div > select:focus {
        border-color: var(--bloomberg-orange) !important;
        outline: none !important;
        box-shadow: 0 0 0 2px rgba(255, 107, 0, 0.2) !important;
    }
    
    /* === BUTTONS - Bloomberg Terminal Style === */
    .stButton > button {
        background: var(--bg-secondary) !important;
        color: var(--text-primary) !important;
        border: 1px solid var(--bloomberg-orange) !important;
        border-radius: 2px !important;
        padding: 8px 20px !important;
        font-weight: 600 !important;
        font-size: var(--font-size-sm) !important;
        font-family: var(--font-mono) !important;
        text-transform: uppercase !important;
        letter-spacing: 0.5px !important;
        transition: all 0.2s ease !important;
        cursor: pointer !important;
    }
    
    .stButton > button:hover {
        background: var(--bloomberg-orange) !important;
        color: var(--bg-primary) !important;
        border-color: var(--bloomberg-orange-hover) !important;
        transform: translateY(-1px) !important;
    }
    
    /* Primary Button */
    .stButton > button[kind="primary"] {
        background: var(--bloomberg-orange) !important;
        color: var(--bg-primary) !important;
        border-color: var(--bloomberg-orange) !important;
    }
    
    .stButton > button[kind="primary"]:hover {
        background: var(--bloomberg-orange-hover) !important;
        border-color: var(--bloomberg-orange-hover) !important;
    }
    
    /* === TABS - Bloomberg Style === */
    .stTabs [data-baseweb="tab-list"] {
        background: var(--bg-secondary) !important;
        border-bottom: 2px solid var(--border-primary) !important;
        gap: 0 !important;
    }
    
    .stTabs [data-baseweb="tab"] {
        background: transparent !important;
        border: none !important;
        color: var(--text-secondary) !important;
        font-family: var(--font-mono) !important;
        font-size: var(--font-size-sm) !important;
        font-weight: 600 !important;
        text-transform: uppercase !important;
        letter-spacing: 0.5px !important;
        padding: 12px 24px !important;
        border-bottom: 3px solid transparent !important;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background: var(--bg-hover) !important;
        color: var(--text-primary) !important;
    }
    
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background: transparent !important;
        color: var(--bloomberg-orange) !important;
        border-bottom-color: var(--bloomberg-orange) !important;
    }
    
    .stTabs [data-baseweb="tab-panel"] {
        background: transparent !important;
    }
    
    /* === METRICS - Bloomberg Style === */
    [data-testid="stMetric"] {
        background: var(--bg-secondary) !important;
        padding: 12px 16px !important;
        border: 1px solid var(--border-primary) !important;
        border-left: 3px solid var(--bloomberg-orange) !important;
        border-radius: 0 !important;
    }
    
    [data-testid="stMetric"]:hover {
        background: var(--bg-hover) !important;
        border-left-color: var(--bloomberg-orange-hover) !important;
    }
    
    [data-testid="stMetricValue"] {
        font-size: 1.8rem !important;
        font-weight: 700 !important;
        color: var(--text-primary) !important;
        font-family: var(--font-mono) !important;
    }
    
    [data-testid="stMetricLabel"] {
        color: var(--text-secondary) !important;
        text-transform: uppercase !important;
        font-size: var(--font-size-xs) !important;
        letter-spacing: 0.5px !important;
        font-weight: 600 !important;
    }
    
    [data-testid="stMetricDelta"] {
        font-family: var(--font-mono) !important;
    }
    
    /* === DATAFRAMES & TABLES - Bloomberg Terminal Style === */
    [data-testid="stDataFrame"],
    [data-testid="stTable"],
    .dataframe {
        background: var(--bg-secondary) !important;
        border: 1px solid var(--border-primary) !important;
        border-radius: 0 !important;
        font-family: var(--font-mono) !important;
        font-size: var(--font-size-xs) !important;
        color: var(--text-primary) !important;
    }
    
    .dataframe thead th {
        background: var(--bg-primary) !important;
        color: var(--bloomberg-orange) !important;
        font-weight: 700 !important;
        padding: 10px 12px !important;
        text-align: left !important;
        border-bottom: 2px solid var(--bloomberg-orange) !important;
        text-transform: uppercase !important;
        letter-spacing: 0.5px !important;
        font-size: var(--font-size-xs) !important;
    }
    
    .dataframe tbody td {
        padding: 8px 12px !important;
        border-bottom: 1px solid var(--border-primary) !important;
        color: var(--text-primary) !important;
        background: var(--bg-secondary) !important;
    }
    
    .dataframe tbody tr:hover {
        background: var(--bg-hover) !important;
    }
    
    .dataframe tbody tr:nth-child(even) {
        background: var(--bg-secondary) !important;
    }
    
    .dataframe tbody tr:nth-child(odd) {
        background: var(--bg-primary) !important;
    }
    
    /* === TEXT & PARAGRAPHS === */
    p, span {
        color: var(--text-primary) !important;
        font-family: var(--font-mono) !important;
    }
    
    /* === PROGRESS BARS === */
    .stProgress > div > div > div {
        background: var(--bloomberg-orange) !important;
    }
    
    /* === INFO/SUCCESS/WARNING/ERROR MESSAGES === */
    [data-baseweb="notification"] {
        border-radius: 0 !important;
        border-left: 3px solid !important;
        background: var(--bg-secondary) !important;
        font-family: var(--font-mono) !important;
        color: var(--text-primary) !important;
    }
    
    /* === SCROLLBARS === */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    
    ::-webkit-scrollbar-track {
        background: var(--bg-primary);
    }
    
    ::-webkit-scrollbar-thumb {
        background: var(--border-secondary);
        border-radius: 0;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: var(--bloomberg-orange);
    }
    
    /* === LINKS === */
    a {
        color: var(--bloomberg-blue) !important;
        text-decoration: none !important;
    }
    
    a:hover {
        color: var(--bloomberg-orange) !important;
    }
    
    [data-testid="stMetric"] {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: var(--radius-md);
        padding: var(--spacing-4);
        box-shadow: var(--shadow-sm);
        transition: all 0.3s ease;
    }
    
    [data-testid="stMetric"]:hover {
        border-color: var(--primary-500);
        box-shadow: var(--shadow-md);
        transform: translateY(-2px);
    }
    
    [data-testid="stMetricLabel"] {
        font-size: var(--font-size-sm) !important;
        font-weight: 600 !important;
        color: var(--text-secondary) !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-bottom: var(--spacing-2);
    }
    
    [data-testid="stMetricDelta"] {
        font-size: var(--font-size-sm) !important;
        font-weight: 500 !important;
    }
    
    /* Checkboxes & Radio */
    .stCheckbox, .stRadio {
        padding: var(--spacing-2) 0;
    }
    
    .stCheckbox > label,
    .stRadio > label {
        color: var(--text-primary);
        font-size: var(--font-size-sm);
        font-weight: 500;
    }
    
    /* Slider */
    .stSlider > div > div > div {
        background: var(--primary-500);
    }
    
    /* Metrics - Enhanced Cards */
    [data-testid="stMetricValue"] {
        font-size: var(--font-size-2xl);
        font-weight: 700;
        color: var(--text-primary);
    }
    
    [data-testid="stMetricDelta"] {
        font-size: var(--font-size-sm);
        font-weight: 600;
    }
    
    div[data-testid="stMetricValue"] > div {
        background: linear-gradient(135deg, var(--primary-600), var(--primary-800));
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    
    /* DataFrame Styling */
    .dataframe {
        border: 1px solid var(--border);
        border-radius: var(--radius);
        overflow: hidden;
        box-shadow: var(--shadow);
        font-size: var(--font-size-sm);
    }
    
    .dataframe thead th {
        background: var(--primary-600);
        color: white;
        font-weight: 700;
        padding: var(--spacing-3) var(--spacing-4);
        text-align: left;
        border-bottom: 2px solid var(--primary-700);
    }
    
    .dataframe tbody td {
        padding: var(--spacing-3) var(--spacing-4);
        border-bottom: 1px solid var(--border);
        color: var(--text-primary);
    }
    
    .dataframe tbody tr:hover {
        background: var(--primary-50);
    }
    
    /* Expander Component */
    .streamlit-expanderHeader {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        padding: var(--spacing-4);
        font-weight: 600;
        color: var(--text-primary);
        transition: all 0.2s ease;
    }
    
    .streamlit-expanderHeader:hover {
        background: var(--primary-50);
        border-color: var(--primary-300);
    }
    
    .streamlit-expanderContent {
        border: 1px solid var(--border);
        border-top: none;
        border-radius: 0 0 var(--radius) var(--radius);
        padding: var(--spacing-4);
        background: var(--surface);
    }
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: var(--spacing-2);
        background: var(--background-secondary);
        padding: var(--spacing-2);
        border-radius: var(--radius);
    }
    
    .stTabs [data-baseweb="tab"] {
        background: transparent;
        border: none;
        color: var(--text-secondary);
        font-weight: 600;
        padding: var(--spacing-3) var(--spacing-6);
        border-radius: var(--radius-sm);
        transition: all 0.2s ease;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background: var(--surface);
        color: var(--primary-700);
    }
    
    .stTabs [aria-selected="true"] {
        background: var(--primary-600);
        color: white;
        box-shadow: var(--shadow);
    }
    
    /* Alert/Info Boxes */
    .stAlert {
        border-radius: var(--radius);
        padding: var(--spacing-4);
        border-left: 4px solid;
        box-shadow: var(--shadow-sm);
    }
    
    .stInfo {
        background: var(--info-bg);
        border-color: var(--info-text);
        color: var(--info-text);
    }
    
    .stSuccess {
        background: var(--success-bg);
        border-color: var(--success-text);
        color: var(--success-text);
    }
    
    .stWarning {
        background: var(--warning-bg);
        border-color: var(--warning-text);
        color: var(--warning-text);
    }
    
    .stError {
        background: var(--error-bg);
        border-color: var(--error-text);
        color: var(--error-text);
    }
    
    /* Progress Bar */
    .stProgress > div > div > div {
        background: var(--primary-600);
        border-radius: var(--radius);
    }
    
    /* Spinner */
    .stSpinner > div {
        border-color: var(--primary-600);
    }
    
    /* Code Blocks */
    code {
        background: var(--neutral-100);
        color: var(--primary-700);
        padding: var(--spacing-1) var(--spacing-2);
        border-radius: var(--radius-sm);
        font-size: var(--font-size-sm);
        font-family: 'Fira Code', monospace;
    }
    
    pre {
        background: var(--neutral-900);
        color: var(--neutral-100);
        padding: var(--spacing-4);
        border-radius: var(--radius);
        overflow-x: auto;
        box-shadow: var(--shadow);
    }
    
    /* Custom Card Component */
    .card {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: var(--radius-lg);
        padding: var(--spacing-6);
        box-shadow: var(--shadow);
        transition: all 0.3s ease;
        margin-bottom: var(--spacing-4);
    }
    
    .card:hover {
        box-shadow: var(--shadow-lg);
        transform: translateY(-2px);
        border-color: var(--primary-300);
    }
    
    .card-header {
        font-size: var(--font-size-xl);
        font-weight: 700;
        color: var(--primary-700);
        margin-bottom: var(--spacing-4);
        padding-bottom: var(--spacing-3);
        border-bottom: 2px solid var(--border);
    }
    
    /* Badge Component */
    .badge {
        display: inline-block;
        padding: var(--spacing-1) var(--spacing-3);
        border-radius: var(--radius);
        font-size: var(--font-size-xs);
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
    .badge-primary {
        background: var(--primary-100);
        color: var(--primary-700);
        border: 1px solid var(--primary-300);
    }
    
    .badge-success {
        background: var(--success-bg);
        color: var(--success-text);
        border: 1px solid var(--success-border);
    }
    
    .badge-warning {
        background: var(--warning-bg);
        color: var(--warning-text);
        border: 1px solid var(--warning-border);
    }
    
    .badge-error {
        background: var(--error-bg);
        color: var(--error-text);
        border: 1px solid var(--error-border);
    }
    
    /* Scrollbar Styling */
    ::-webkit-scrollbar {
        width: 10px;
        height: 10px;
    }
    
    ::-webkit-scrollbar-track {
        background: var(--background-secondary);
        border-radius: var(--radius);
    }
    
    ::-webkit-scrollbar-thumb {
        background: var(--primary-400);
        border-radius: var(--radius);
        transition: all 0.2s ease;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: var(--primary-600);
    }
    
    /* Loading Animation */
    @keyframes pulse {
        0%, 100% {
            opacity: 1;
        }
        50% {
            opacity: 0.5;
        }
    }
    
    .loading {
        animation: pulse 2s cubic-bezier(0.4, 0, 0.6, 1) infinite;
    }
    
    /* Fade In Animation */
    @keyframes fadeIn {
        from {
            opacity: 0;
            transform: translateY(10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .fade-in {
        animation: fadeIn 0.5s ease-out;
    }
    
    /* Responsive Design - Mobile & Tablet */
    @media (max-width: 768px) {
        /* Typography adjustments */
        h1 {
            font-size: 1.5rem;
        }
        
        h2 {
            font-size: 1.2rem;
        }
        
        h3 {
            font-size: 1rem;
        }
        
        /* Main content padding */
        .main {
            padding: var(--spacing-3) var(--spacing-2);
        }
        
        /* Card padding reduction */
        .card {
            padding: var(--spacing-3);
        }
        
        /* Sidebar responsive */
        [data-testid="stSidebar"] {
            min-width: 280px !important;
            max-width: 280px !important;
        }
        
        /* Collapsible sidebar button always visible */
        [data-testid="collapsedControl"] {
            display: block !important;
            background: var(--bloomberg-orange) !important;
            color: white !important;
            border-radius: 0 8px 8px 0 !important;
            padding: 8px !important;
        }
        
        /* Table/DataFrame responsive */
        [data-testid="stDataFrame"],
        [data-testid="stTable"] {
            font-size: 0.75rem !important;
            overflow-x: auto !important;
        }
        
        /* Metric cards stack better */
        [data-testid="stMetric"] {
            margin-bottom: 8px !important;
        }
        
        /* Chart height adjustment */
        .plotly {
            height: 400px !important;
        }
    }
    
    /* Tablet specific (768px - 1024px) */
    @media (min-width: 769px) and (max-width: 1024px) {
        [data-testid="stSidebar"] {
            min-width: 320px !important;
            max-width: 320px !important;
        }
        
        .main {
            padding: var(--spacing-5);
        }
    }
    
    /* Desktop - Sidebar collapse control */
    @media (min-width: 1025px) {
        /* Show collapse button on desktop too for flexibility */
        [data-testid="collapsedControl"] {
            display: block !important;
            background: rgba(255, 107, 0, 0.1) !important;
            border: 1px solid var(--bloomberg-orange) !important;
            transition: all 0.3s ease !important;
        }
        
        [data-testid="collapsedControl"]:hover {
            background: var(--bloomberg-orange) !important;
            color: white !important;
        }
    }
    
    /* Print Styles */
    @media print {
        .stButton, .stDownloadButton, [data-testid="stSidebar"] {
            display: none;
        }
        
        .main {
            padding: 0;
        }
        
        .card {
            box-shadow: none;
            border: 1px solid var(--border);
            page-break-inside: avoid;
        }
    }
    
    /* High Contrast Mode Support */
    @media (prefers-contrast: high) {
        :root {
            --border: hsl(210, 14%, 70%);
            --text-secondary: hsl(210, 12%, 30%);
        }
    }
    
    /* Reduced Motion Support */
    @media (prefers-reduced-motion: reduce) {
        * {
            animation-duration: 0.01ms !important;
            animation-iteration-count: 1 !important;
            transition-duration: 0.01ms !important;
        }
    }
    
    /* Focus Visible for Accessibility */
    *:focus-visible {
        outline: 2px solid var(--primary-600);
        outline-offset: 2px;
        border-radius: var(--radius-sm);
    }
    
    /* Selection Styling */
    ::selection {
        background: var(--primary-200);
        color: var(--primary-900);
    }
    
    /* Custom Professional Elements */
    .pro-header {
        background: linear-gradient(135deg, var(--primary-600), var(--primary-800));
        color: white;
        padding: var(--spacing-8);
        border-radius: var(--radius-lg);
        margin-bottom: var(--spacing-6);
        box-shadow: var(--shadow-xl);
    }
    
    .pro-stat {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        padding: var(--spacing-4);
        text-align: center;
        transition: all 0.3s ease;
    }
    
    .pro-stat:hover {
        border-color: var(--primary-500);
        box-shadow: var(--shadow-md);
        transform: scale(1.02);
    }
    
    .pro-stat-value {
        font-size: var(--font-size-3xl);
        font-weight: 800;
        color: var(--primary-700);
        display: block;
        margin-bottom: var(--spacing-2);
    }
    
    .pro-stat-label {
        font-size: var(--font-size-sm);
        color: var(--text-secondary);
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
</style>
""", unsafe_allow_html=True)

# COMPLETE NSE F&O UNIVERSE - 208 STOCKS (Updated Official List)
COMPLETE_NSE_FO_UNIVERSE = [
    # Updated 208 F&O Individual stocks - As per user's official list
    '360ONE.NS',
    'ABB.NS',
    'APLAPOLLO.NS',
    'AUBANK.NS',
    'ADANIENSOL.NS',
    'ADANIENT.NS',
    'ADANIGREEN.NS',
    'ADANIPORTS.NS',
    'ABCAPITAL.NS',
    'ALKEM.NS',
    'AMBER.NS',
    'AMBUJACEM.NS',
    'ANGELONE.NS',
    'APOLLOHOSP.NS',
    'ASHOKLEY.NS',
    'ASIANPAINT.NS',
    'ASTRAL.NS',
    'AUROPHARMA.NS',
    'DMART.NS',
    'AXISBANK.NS',
    'BSE.NS',
    'BAJAJ-AUTO.NS',
    'BAJFINANCE.NS',
    'BAJAJFINSV.NS',
    'BAJAJHLDNG.NS',
    'BANDHANBNK.NS',
    'BANKBARODA.NS',
    'BANKINDIA.NS',
    'BDL.NS',
    'BEL.NS',
    'BHARATFORG.NS',
    'BHEL.NS',
    'BPCL.NS',
    'BHARТIARTL.NS',
    'BIOCON.NS',
    'BLUESTARCO.NS',
    'BOSCHLTD.NS',
    'BRITANNIA.NS',
    'CGPOWER.NS',
    'CANBK.NS',
    'CDSL.NS',
    'CHOLAFIN.NS',
    'CIPLA.NS',
    'COALINDIA.NS',
    'COFORGE.NS',
    'COLPAL.NS',
    'CAMS.NS',
    'CONCOR.NS',
    'CROMPTON.NS',
    'CUMMINSIND.NS',
    'DLF.NS',
    'DABUR.NS',
    'DALBHARAT.NS',
    'DELHIVERY.NS',
    'DIVISLAB.NS',
    'DIXON.NS',
    'DRREDDY.NS',
    'ETERNAL.NS',
    'EICHERMOT.NS',
    'EXIDEIND.NS',
    'NYKAA.NS',
    'FORTIS.NS',
    'GAIL.NS',
    'GMRAIRPORT.NS',
    'GLENMARK.NS',
    'GODREJCP.NS',
    'GODREJPROP.NS',
    'GRASIM.NS',
    'HCLTECH.NS',
    'HDFCAMC.NS',
    'HDFCBANK.NS',
    'HDFCLIFE.NS',
    'HAVELLS.NS',
    'HEROMOTOCO.NS',
    'HINDALCO.NS',
    'HAL.NS',
    'HINDPETRO.NS',
    'HINDUNILVR.NS',
    'HINDZINC.NS',
    'POWERINDIA.NS',
    'HUDCO.NS',
    'ICICIBANK.NS',
    'ICICIGI.NS',
    'ICICIPRULI.NS',
    'IDFCFIRSTB.NS',
    'IIFL.NS',
    'ITC.NS',
    'INDIANB.NS',
    'IEX.NS',
    'IOC.NS',
    'IRCTC.NS',
    'IRFC.NS',
    'IREDA.NS',
    'INDUSTOWER.NS',
    'INDUSINDBK.NS',
    'NAUKRI.NS',
    'INFY.NS',
    'INOXWIND.NS',
    'INDIGO.NS',
    'JINDALSTEL.NS',
    'JSWENERGY.NS',
    'JSWSTEEL.NS',
    'JIOFIN.NS',
    'JUBLFOOD.NS',
    'KEI.NS',
    'KPITTECH.NS',
    'KALYANKJIL.NS',
    'KAYNES.NS',
    'KFINTECH.NS',
    'KOTAKBANK.NS',
    'LTF.NS',
    'LICHSGFIN.NS',
    'LTIM.NS',
    'LT.NS',
    'LAURUSLABS.NS',
    'LICI.NS',
    'LODHA.NS',
    'LUPIN.NS',
    'M&M.NS',
    'MANAPPURAM.NS',
    'MANKIND.NS',
    'MARICO.NS',
    'MARUTI.NS',
    'MFSL.NS',
    'MAXHEALTH.NS',
    'MAZDOCK.NS',
    'MPHASIS.NS',
    'MCX.NS',
    'MUTHOOTFIN.NS',
    'NBCC.NS',
    'NHPC.NS',
    'NMDC.NS',
    'NTPC.NS',
    'NATIONALUM.NS',
    'NESTLEIND.NS',
    'NUVAMA.NS',
    'OBEROIRLTY.NS',
    'ONGC.NS',
    'OIL.NS',
    'PAYTM.NS',
    'OFSS.NS',
    'POLICYBZR.NS',
    'PGEL.NS',
    'PIIND.NS',
    'PNBHOUSING.NS',
    'PAGEIND.NS',
    'PATANJALI.NS',
    'PERSISTENT.NS',
    'PETRONET.NS',
    'PIDILITIND.NS',
    'PPLPHARMA.NS',
    'POLYCAB.NS',
    'PFC.NS',
    'POWERGRID.NS',
    'PREMIERENE.NS',
    'PRESTIGE.NS',
    'PNB.NS',
    'RBLBANK.NS',
    'RECLTD.NS',
    'RVNL.NS',
    'RELIANCE.NS',
    'SBICARD.NS',
    'SBILIFE.NS',
    'SHREECEM.NS',
    'SRF.NS',
    'SAMMAANCAP.NS',
    'MOTHERSON.NS',
    'SHRIRAMFIN.NS',
    'SIEMENS.NS',
    'SOLARINDS.NS',
    'SONACOMS.NS',
    'SBIN.NS',
    'SAIL.NS',
    'SUNPHARMA.NS',
    'SUPREMEIND.NS',
    'SUZLON.NS',
    'SWIGGY.NS',
    'SYNGENE.NS',
    'TATACONSUM.NS',
    'TVSMOTOR.NS',
    'TCS.NS',
    'TATAELXSI.NS',
    'TMPV.NS',
    'TATAPOWER.NS',
    'TATASTEEL.NS',
    'TATATECH.NS',
    'TECHM.NS',
    'FEDERALBNK.NS',
    'INDHOTEL.NS',
    'PHOENIXLTD.NS',
    'TITAN.NS',
    'TORNTPHARM.NS',
    'TORNTPOWER.NS',
    'TRENT.NS',
    'TIINDIA.NS',
    'UNOMINDA.NS',
    'UPL.NS',
    'ULTRACEMCO.NS',
    'UNIONBANK.NS',
    'UNITDSPR.NS',
    'VBL.NS',
    'VEDL.NS',
    'IDEA.NS',
    'VOLTAS.NS',
    'WAAREEENER.NS',
    'WIPRO.NS',
    'YESBANK.NS',
    'ZYDUSLIFE.NS'
]

# Stock categories with verified symbols
STOCK_CATEGORIES = {
    'Nifty 50': [
        'RELIANCE.NS', 'TCS.NS', 'HDFCBANK.NS', 'INFY.NS', 'ICICIBANK.NS',
        'BHARTIARTL.NS', 'ITC.NS', 'SBIN.NS', 'LT.NS', 'KOTAKBANK.NS',
        'AXISBANK.NS', 'MARUTI.NS', 'ASIANPAINT.NS', 'WIPRO.NS', 'ONGC.NS',
        'NTPC.NS', 'POWERGRID.NS', 'TATAMOTORS.NS', 'TECHM.NS', 'ULTRACEMCO.NS',
        'SUNPHARMA.NS', 'TITAN.NS', 'COALINDIA.NS', 'BAJFINANCE.NS', 'HCLTECH.NS',
        'JSWSTEEL.NS', 'INDUSINDBK.NS', 'BRITANNIA.NS', 'CIPLA.NS', 'DRREDDY.NS',
        'EICHERMOT.NS', 'GRASIM.NS', 'HEROMOTOCO.NS', 'HINDALCO.NS', 'TATASTEEL.NS',
        'BPCL.NS', 'M&M.NS', 'BAJAJ-AUTO.NS', 'SHRIRAMFIN.NS', 'ADANIPORTS.NS',
        'APOLLOHOSP.NS', 'BAJAJFINSV.NS', 'DIVISLAB.NS', 'NESTLEIND.NS', 'TRENT.NS',
        'HDFCLIFE.NS', 'SBILIFE.NS', 'LTIM.NS', 'ADANIENT.NS', 'HINDUNILVR.NS'
    ],
    'Bank Nifty': [
        'HDFCBANK.NS', 'ICICIBANK.NS', 'SBIN.NS', 'KOTAKBANK.NS', 'AXISBANK.NS',
        'INDUSINDBK.NS', 'BANKBARODA.NS', 'CANBK.NS', 'FEDERALBNK.NS', 'PNB.NS',
        'IDFCFIRSTB.NS', 'AUBANK.NS'
    ],
    'IT Stocks': [
        'TCS.NS', 'INFY.NS', 'WIPRO.NS', 'HCLTECH.NS', 'TECHM.NS',
        'LTIM.NS', 'MPHASIS.NS', 'COFORGE.NS', 'PERSISTENT.NS', 'LTTS.NS'
    ],
    'Pharma Stocks': [
        'SUNPHARMA.NS', 'CIPLA.NS', 'DRREDDY.NS', 'DIVISLAB.NS', 'LUPIN.NS',
        'BIOCON.NS', 'AUROPHARMA.NS', 'ALKEM.NS', 'TORNTPHARM.NS', 'GLENMARK.NS'
    ],
    'Auto Stocks': [
        'MARUTI.NS', 'TATAMOTORS.NS', 'M&M.NS', 'BAJAJ-AUTO.NS', 'HEROMOTOCO.NS',
        'EICHERMOT.NS', 'TVSMOTOR.NS', 'ASHOKLEY.NS', 'ESCORTS.NS'
    ],
    'Metal Stocks': [
        'TATASTEEL.NS', 'JSWSTEEL.NS', 'HINDALCO.NS', 'COALINDIA.NS', 'VEDL.NS',
        'JINDALSTEL.NS', 'NATIONALUM.NS', 'NMDC.NS', 'SAIL.NS', 'HINDZINC.NS'
    ],
    'Energy Stocks': [
        'RELIANCE.NS', 'ONGC.NS', 'IOC.NS', 'BPCL.NS', 'HINDPETRO.NS',
        'GAIL.NS', 'NTPC.NS', 'POWERGRID.NS', 'TATAPOWER.NS', 'ADANIGREEN.NS'
    ]
}

def get_nse_non_fno_stocks():
    """
    Fetch all NSE stocks and exclude F&O stocks to get non-F&O universe.
    Returns list of stock symbols with .NS suffix.
    """
    try:
        # Try fetching from NSE website
        url = "https://archives.nseindia.com/content/equities/EQUITY_L.csv"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
        }
        
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            from io import StringIO
            df = pd.read_csv(StringIO(response.text))
            
            # Get all symbols and add .NS suffix
            all_symbols = [f"{symbol.strip()}.NS" for symbol in df['SYMBOL'].tolist() if pd.notna(symbol)]
            
            # Remove F&O stocks
            fno_symbols_clean = [s.replace('.NS', '') for s in COMPLETE_NSE_FO_UNIVERSE if s.endswith('.NS')]
            non_fno_stocks = [s for s in all_symbols if s.replace('.NS', '') not in fno_symbols_clean]
            
            return non_fno_stocks
        else:
            return _get_comprehensive_backup_list()
            
    except Exception:
        return _get_comprehensive_backup_list()

def create_excel_stock_list(results):
    """
    Create Excel file with just stock symbols that met criteria.
    Simple format: One column with stock symbols.
    """
    # Extract clean symbols from results
    stock_symbols = []
    for result in results:
        # Remove .NS suffix and clean symbol
        clean_symbol = result['symbol'].replace('.NS', '').replace('^', '')
        stock_symbols.append(clean_symbol)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Qualifying Stocks"
    
    # Add header
    ws['A1'] = "Stock Symbol"
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=12)
    
    # Add stock symbols
    for idx, symbol in enumerate(stock_symbols, start=2):
        ws[f'A{idx}'] = symbol
    
    # Auto-adjust column width
    ws.column_dimensions['A'].width = 20
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def _get_comprehensive_backup_list():
    """Comprehensive backup list of 814+ NSE stocks (non-F&O)"""
    return [
        # (backup list contents omitted here for brevity – leave as in your original file)
    ]


# Cached data fetching for better performance
@st.cache_data(ttl=300)  # Cache for 5 minutes
def fetch_stock_data_cached(symbol, period="3mo"):
    """Cached stock data fetch to avoid repeated API calls"""
    try:
        stock = yf.Ticker(symbol)
        data = stock.history(period=period, interval="1d")
        return data if len(data) >= 20 else None
    except Exception:
        return None

@st.cache_data(ttl=300)
def fetch_weekly_data_cached(symbol, period="6mo"):
    """Cached weekly stock data fetch"""
    try:
        stock = yf.Ticker(symbol)
        data = stock.history(period=period, interval="1wk")
        return data if len(data) >= 8 else None
    except Exception:
        return None

class ProfessionalPCSScanner:
    def __init__(self):
        self.ist = pytz.timezone('Asia/Kolkata')
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
    
    def get_stock_data(self, symbol, period="3mo"):
        """Get stock data with focus on recent data for current trading day analysis"""
        try:
            data = fetch_stock_data_cached(symbol, period)
            if data is None or len(data) < 20:
                return None
            
            # Technical indicators
            data['RSI'] = ta.momentum.RSIIndicator(data['Close']).rsi()
            data['SMA_20'] = ta.trend.SMAIndicator(data['Close'], window=20).sma_indicator()
            data['SMA_50'] = ta.trend.SMAIndicator(data['Close'], window=50).sma_indicator()
            data['EMA_20'] = ta.trend.EMAIndicator(data['Close'], window=20).ema_indicator()
            bb = ta.volatility.BollingerBands(data['Close'])
            data['BB_upper'] = bb.bollinger_hband()
            data['BB_lower'] = bb.bollinger_lband()
            data['BB_middle'] = bb.bollinger_mavg()
            macd = ta.trend.MACD(data['Close'])
            data['MACD'] = macd.macd()
            data['MACD_signal'] = macd.macd_signal()
            data['MACD_hist'] = macd.macd_diff()
            adx = ta.trend.ADXIndicator(data['High'], data['Low'], data['Close'])
            data['ADX'] = adx.adx()
            atr = ta.volatility.AverageTrueRange(data['High'], data['Low'], data['Close'])
            data['ATR'] = atr.average_true_range()
            stoch = ta.momentum.StochasticOscillator(data['High'], data['Low'], data['Close'])
            data['Stoch_K'] = stoch.stoch()
            will = ta.momentum.WilliamsRIndicator(data['High'], data['Low'], data['Close'])
            data['Williams_R'] = will.williams_r()
            return data
        except Exception:
            return None
    
    def get_weekly_stock_data(self, symbol, period="6mo"):
        """Get weekly stock data for pattern validation"""
        try:
            daily_data = fetch_stock_data_cached(symbol, period)
            if daily_data is None or len(daily_data) < 50:
                return None
            
            weekly_data = daily_data.resample('W-FRI').agg({
                'Open': 'first',
                'High': 'max',
                'Low': 'min',
                'Close': 'last',
                'Volume': 'sum'
            }).dropna()
            
            if len(weekly_data) < 15:
                return None
            
            weekly_data['RSI'] = ta.momentum.RSIIndicator(weekly_data['Close']).rsi()
            weekly_data['SMA_10'] = ta.trend.SMAIndicator(weekly_data['Close'], window=10).sma_indicator()
            weekly_data['SMA_20'] = ta.trend.SMAIndicator(weekly_data['Close'], window=20).sma_indicator()
            weekly_data['EMA_10'] = ta.trend.EMAIndicator(weekly_data['Close'], window=10).ema_indicator()
            macd = ta.trend.MACD(weekly_data['Close'])
            weekly_data['MACD'] = macd.macd()
            weekly_data['MACD_signal'] = macd.macd_signal()
            weekly_data['MACD_hist'] = macd.macd_diff()
            adx = ta.trend.ADXIndicator(weekly_data['High'], weekly_data['Low'], weekly_data['Close'])
            weekly_data['ADX'] = adx.adx()
            return weekly_data
        except Exception:
            return None
    
    def validate_weekly_strength(self, daily_data, weekly_data, pattern_type):
        """Validate daily pattern strength using weekly timeframe analysis"""
        if weekly_data is None or len(weekly_data) < 10:
            return {
                'weekly_validation': False,
                'weekly_strength_bonus': 0,
                'weekly_signals': [],
                'weekly_context': 'Insufficient weekly data'
            }
        
        try:
            current_weekly_close = weekly_data['Close'].iloc[-1]
            current_weekly_rsi = float(weekly_data['RSI'].iloc[-1]) if not pd.isna(weekly_data['RSI'].iloc[-1]) else 50.0
            current_weekly_macd = weekly_data['MACD'].iloc[-1]
            current_weekly_macd_signal = weekly_data['MACD_signal'].iloc[-1]
            weekly_sma_10 = weekly_data['SMA_10'].iloc[-1]
            weekly_sma_20 = weekly_data['SMA_20'].iloc[-1]
            current_weekly_adx = weekly_data['ADX'].iloc[-1]
            
            weekly_signals = []
            strength_bonus = 0
            
            # 1. Weekly Trend Alignment
            if current_weekly_close > weekly_sma_10 > weekly_sma_20:
                weekly_signals.append("Strong weekly uptrend (Close > SMA10 > SMA20)")
                strength_bonus += 15
            elif current_weekly_close > weekly_sma_10:
                weekly_signals.append("Bullish weekly trend (Close > SMA10)")
                strength_bonus += 10
            elif current_weekly_close > weekly_sma_20:
                weekly_signals.append("Weekly above long-term MA")
                strength_bonus += 5
            
            # 2. Weekly RSI Support
            if 40 <= current_weekly_rsi <= 70:
                weekly_signals.append(f"Healthy weekly RSI ({current_weekly_rsi:.1f})")
                strength_bonus += 10
            elif current_weekly_rsi > 30:
                weekly_signals.append(f"Weekly RSI above oversold ({current_weekly_rsi:.1f})")
                strength_bonus += 5
            
            # 3. Weekly MACD Confirmation
            if current_weekly_macd > current_weekly_macd_signal and current_weekly_macd > 0:
                weekly_signals.append("Weekly MACD bullish above signal line")
                strength_bonus += 12
            elif current_weekly_macd > current_weekly_macd_signal:
                weekly_signals.append("Weekly MACD above signal line")
                strength_bonus += 8
            
            # 4. Weekly ADX Trend Strength
            if current_weekly_adx >= 25:
                weekly_signals.append(f"Strong weekly trend (ADX: {current_weekly_adx:.1f})")
                strength_bonus += 8
            elif current_weekly_adx >= 20:
                weekly_signals.append(f"Moderate weekly trend (ADX: {current_weekly_adx:.1f})")
                strength_bonus += 5
            
            # 5. Weekly Support/Resistance Context
            weekly_support_resistance = self._analyze_weekly_support_resistance(weekly_data)
            if weekly_support_resistance['near_breakout']:
                weekly_signals.append(weekly_support_resistance['context'])
                strength_bonus += weekly_support_resistance['bonus']
            
            # 6. Weekly Volume Trend
            weekly_volume_trend = self._analyze_weekly_volume_trend(weekly_data)
            if weekly_volume_trend['positive']:
                weekly_signals.append(weekly_volume_trend['context'])
                strength_bonus += weekly_volume_trend['bonus']
            
            # 7. Pattern-Specific Weekly Validation
            pattern_bonus = self._get_pattern_specific_weekly_bonus(pattern_type, weekly_data)
            if pattern_bonus['bonus'] > 0:
                weekly_signals.append(pattern_bonus['context'])
                strength_bonus += pattern_bonus['bonus']
            
            weekly_validation = len(weekly_signals) >= 2 and strength_bonus >= 15
            
            if strength_bonus >= 35:
                weekly_context = "Exceptionally strong weekly confirmation"
            elif strength_bonus >= 25:
                weekly_context = "Strong weekly alignment"
            elif strength_bonus >= 15:
                weekly_context = "Moderate weekly support"
            else:
                weekly_context = "Weak weekly confirmation"
            
            return {
                'weekly_validation': weekly_validation,
                'weekly_strength_bonus': strength_bonus,
                'weekly_signals': weekly_signals,
                'weekly_context': weekly_context,
                'weekly_rsi': current_weekly_rsi,
                'weekly_trend': 'Bullish' if current_weekly_close > weekly_sma_10 else 'Neutral/Bearish'
            }
            
        except Exception as e:
            return {
                'weekly_validation': False,
                'weekly_strength_bonus': 0,
                'weekly_signals': [],
                'weekly_context': f'Weekly analysis error: {str(e)}'
            }
    
    def _analyze_weekly_support_resistance(self, weekly_data):
        """Analyze weekly support/resistance levels"""
        try:
            recent_weeks = weekly_data.tail(12)
            current_price = weekly_data['Close'].iloc[-1]
            
            resistance_level = recent_weeks['High'].max()
            support_level = recent_weeks['Low'].min()
            
            distance_to_resistance = ((resistance_level - current_price) / current_price) * 100
            distance_from_support = ((current_price - support_level) / support_level) * 100
            
            if distance_to_resistance <= 3:
                return {
                    'near_breakout': True,
                    'context': f"Near weekly resistance breakout (~{distance_to_resistance:.1f}% away)",
                    'bonus': 12
                }
            elif distance_from_support >= 15:
                return {
                    'near_breakout': True,
                    'context': f"Strong weekly support base ({distance_from_support:.1f}% above support)",
                    'bonus': 8
                }
            
            return {'near_breakout': False, 'context': '', 'bonus': 0}
        except Exception:
            return {'near_breakout': False, 'context': '', 'bonus': 0}
    
    def _analyze_weekly_volume_trend(self, weekly_data):
        """Analyze weekly volume trends"""
        try:
            recent_volume = weekly_data['Volume'].tail(4).mean()
            previous_volume = weekly_data['Volume'].tail(8).iloc[:4].mean()
            if previous_volume == 0 or pd.isna(previous_volume):
                return {'positive': False, 'context': '', 'bonus': 0}
            
            volume_increase = ((recent_volume - previous_volume) / previous_volume) * 100
            
            if volume_increase >= 20:
                return {
                    'positive': True,
                    'context': f"Strong weekly volume increase ({volume_increase:.1f}%)",
                    'bonus': 10
                }
            elif volume_increase >= 10:
                return {
                    'positive': True,
                    'context': f"Moderate weekly volume increase ({volume_increase:.1f}%)",
                    'bonus': 6
                }
            return {'positive': False, 'context': '', 'bonus': 0}
        except Exception:
            return {'positive': False, 'context': '', 'bonus': 0}
    
    def _get_pattern_specific_weekly_bonus(self, pattern_type, weekly_data):
        """Get pattern-specific weekly validation bonus"""
        try:
            if pattern_type in ['Cup and Handle', 'Double Bottom (Eve & Eve)', 'Head-and-Shoulders Bottom']:
                recent_weeks = weekly_data.tail(8)
                price_min = recent_weeks['Low'].min()
                price_max = recent_weeks['High'].max()
                if price_min <= 0:
                    return {'bonus': 0, 'context': ''}
                weekly_consolidation_range = ((price_max - price_min) / price_min) * 100
                
                if weekly_consolidation_range < 20:
                    return {
                        'bonus': 10,
                        'context': f"Tight weekly consolidation ({weekly_consolidation_range:.1f}% range)"
                    }
            
            elif pattern_type in ['Current Day Breakout', 'Rectangle Bottom', 'Flat Base Breakout']:
                if len(weekly_data) <= 5:
                    return {'bonus': 0, 'context': ''}
                current_week = weekly_data['Close'].iloc[-1]
                four_weeks_ago = weekly_data['Close'].iloc[-5]
                if four_weeks_ago <= 0:
                    return {'bonus': 0, 'context': ''}
                weekly_momentum = ((current_week - four_weeks_ago) / four_weeks_ago) * 100
                
                if weekly_momentum > 5:
                    return {
                        'bonus': 12,
                        'context': f"Strong weekly momentum ({weekly_momentum:.1f}% over 4 weeks)"
                    }
                elif weekly_momentum > 0:
                    return {
                        'bonus': 6,
                        'context': f"Positive weekly momentum ({weekly_momentum:.1f}% over 4 weeks)"
                    }
            return {'bonus': 0, 'context': ''}
        except Exception:
            return {'bonus': 0, 'context': ''}
    
    def check_volume_criteria(self, data, min_ratio=1.0):
        """Check volume criteria with focus on latest trading day"""
        if len(data) < 21:
            return False, 0, {}
        
        current_volume = data['Volume'].iloc[-1]
        avg_5_volume = data['Volume'].tail(6).iloc[:-1].mean()
        avg_10_volume = data['Volume'].tail(11).iloc[:-1].mean()
        avg_20_volume = data['Volume'].tail(21).iloc[:-1].mean()
        if any(v <= 0 or pd.isna(v) for v in [avg_5_volume, avg_10_volume, avg_20_volume]):
            return False, 0, {}
        
        volume_ratio_20 = current_volume / avg_20_volume
        volume_ratio_10 = current_volume / avg_10_volume
        volume_ratio_5 = current_volume / avg_5_volume
        
        details = {
            'current_volume': current_volume,
            'avg_5_volume': avg_5_volume,
            'avg_10_volume': avg_10_volume,
            'avg_20_volume': avg_20_volume,
            'ratio_5d': volume_ratio_5,
            'ratio_10d': volume_ratio_10,
            'ratio_20d': volume_ratio_20
        }
        
        return volume_ratio_20 >= min_ratio, volume_ratio_20, details
    
    def get_fundamental_news(self, symbol, stock_name):
        """Get fundamental news for the stock to explain volume/price movements"""
        try:
            clean_symbol = symbol.replace('.NS', '')
            search_queries = [
                f"{stock_name} stock news today",
                f"{clean_symbol} earnings results latest",
                f"{stock_name} order announcement recent"
            ]
            
            news_items = []
            
            for query in search_queries[:2]:
                try:
                    search_url = f"https://www.google.com/search?q={query}&tbm=nws&tbs=qdr:d"
                    response = self.session.get(search_url, timeout=3)
                    if response.status_code == 200:
                        soup = BeautifulSoup(response.content, 'html.parser')
                        news_elements = soup.find_all('div', class_='BNeawe vvjwJb AP7Wnd')[:2]
                        for element in news_elements:
                            headline = element.get_text().strip()
                            if len(headline) > 20:
                                news_items.append({
                                    'headline': headline,
                                    'relevance': self._assess_news_relevance(headline),
                                    'source': 'Recent News'
                                })
                except Exception:
                    continue
            
            if news_items:
                positive_keywords = ['order', 'win', 'contract', 'growth', 'profit', 'beat', 'strong', 'positive', 'approval']
                negative_keywords = ['loss', 'decline', 'weak', 'concern', 'fall', 'drop', 'negative', 'warning']
                
                sentiment_score = 0
                for item in news_items:
                    headline_lower = item['headline'].lower()
                    for word in positive_keywords:
                        if word in headline_lower:
                            sentiment_score += 1
                    for word in negative_keywords:
                        if word in headline_lower:
                            sentiment_score -= 1
                
                overall_sentiment = 'positive' if sentiment_score > 0 else 'negative' if sentiment_score < 0 else 'neutral'
            else:
                overall_sentiment = 'neutral'
            
            return {
                'news_items': news_items[:2],
                'overall_sentiment': overall_sentiment,
                'news_count': len(news_items)
            }
        except Exception:
            return {
                'news_items': [],
                'overall_sentiment': 'neutral',
                'news_count': 0
            }
    
    def _assess_news_relevance(self, headline):
        """Assess how relevant news is to volume/price movement"""
        high_relevance_keywords = ['order', 'contract', 'earnings', 'results', 'approval', 'launch', 'merger']
        medium_relevance_keywords = ['growth', 'expansion', 'investment', 'partnership', 'policy']
        
        headline_lower = headline.lower()
        
        for word in high_relevance_keywords:
            if word in headline_lower:
                return 'high'
        
        for word in medium_relevance_keywords:
            if word in headline_lower:
                return 'medium'
        
        return 'low'
    
    # (All pattern detection, S/R, and enhancement methods remain as in your version above.
    # I’m keeping them unchanged here to avoid another massive block of repeated code.
    # Functionally, nothing changes for metrics vs. your latest code you uploaded.)
    #
    # ------------- SNIP -------------
    # Paste the rest of your class methods here unchanged.
    # ------------- SNIP -------------


def create_professional_sidebar():
    """Create professional sidebar with Angel One styling"""
    with st.sidebar:
        st.markdown("### 📊 Stock Universe")
        
        universe_option = st.radio(
            "Select Stock Universe:",
            ["NSE F&O Stocks (219)", "NSE Non-F&O Stocks (800+)"],
            help="Toggle between F&O and Non-F&O stock universes",
            index=0
        )
        
        if universe_option == "NSE F&O Stocks (219)":
            stocks_to_scan = COMPLETE_NSE_FO_UNIVERSE
            st.success(f"🎯 **F&O Universe**: {len(stocks_to_scan)} stocks with futures & options")
        else:
            with st.spinner("🔄 Fetching NSE Non-F&O stocks..."):
                stocks_to_scan = get_nse_non_fno_stocks()
            st.success(f"📈 **Non-F&O Universe**: {len(stocks_to_scan)} liquid NSE stocks")
        
        st.markdown("### ⚙️ Core Filters")
        
        with st.expander("🎯 Technical Settings", expanded=True):
            col1, col2 = st.columns(2)
            with col1:
                rsi_min = st.slider("RSI Min:", 20, 80, 30)
            with col2:
                rsi_max = st.slider("RSI Max:", 20, 80, 75)
            
            adx_min = st.slider("ADX Minimum:", 10, 50, 20)
            
            ma_support = st.checkbox("Moving Average Support", value=True)
            if ma_support:
                col1, col2 = st.columns(2)
                with col1:
                    ma_type = st.selectbox("MA Type:", ["EMA", "SMA"])
                with col2:
                    ma_tolerance = st.slider("MA Tolerance %:", 0, 10, 3)
            else:
                ma_type = 'EMA'
                ma_tolerance = 3
        
        with st.expander("📊 Volume & Breakout", expanded=True):
            min_volume_ratio = st.slider("Min Volume Ratio:", 0.8, 5.0, 1.2, 0.1)
            volume_breakout_ratio = st.slider("Breakout Volume:", 1.5, 5.0, 2.0, 0.1)
            lookback_days = st.slider("Lookback Period:", 15, 30, 20)
        
        st.markdown("### 📈 Chart Pattern Filters")
        with st.expander("🎯 Pattern Selection", expanded=False):
            st.markdown("**Select patterns to detect:**")
            col1, col2 = st.columns(2)
            with col1:
                pattern_filters = {
                    'current_day_breakout': st.checkbox("Current Day Breakout", value=True),
                    'cup_and_handle': st.checkbox("Cup with Handle", value=True),
                    'flat_base': st.checkbox("Flat Base Breakout", value=True),
                    'bump_and_run': st.checkbox("Bump-and-Run Reversal", value=True),
                    'rectangle_bottom': st.checkbox("Rectangle Bottom", value=True),
                    'rectangle_top': st.checkbox("Rectangle Top", value=True),
                }
            with col2:
                pattern_filters.update({
                    'head_shoulders_bottom': st.checkbox("Head-and-Shoulders Bottom", value=True),
                    'double_bottom': st.checkbox("Double Bottom (Eve & Eve)", value=True),
                    'three_rising_valleys': st.checkbox("Three Rising Valleys", value=True),
                    'rounding_bottom': st.checkbox("Rounding Bottom", value=True),
                    'rounding_top_upside': st.checkbox("Rounding Top (Upside Break)", value=True),
                    'inverted_scallop': st.checkbox("Inverted Scallop", value=True),
                })
            
            pattern_priority = st.radio(
                "Choose detection approach:",
                ["All Patterns (Comprehensive)", "High Success Rate Only (>80%)", "PCS Optimized (>90% suitability)"],
                index=0
            )
            
            st.markdown("**🔍 Timeframe Analysis:**")
            analysis_mode = st.radio(
                "Select Analysis Mode:",
                [
                    "Daily Only (V6.0 Style)",
                    "Weekly Only (New Feature)",
                    "Daily + Weekly Combined (Recommended)"
                ],
                index=2
            )
            
            if analysis_mode == "Daily Only (V6.0 Style)":
                st.info("📊 Daily Analysis only (fast, no weekly confirmation).")
                enable_daily_analysis = True
                enable_weekly_validation = False
            elif analysis_mode == "Weekly Only (New Feature)":
                st.info("📈 Weekly Analysis only (longer-term patterns).")
                enable_daily_analysis = False
                enable_weekly_validation = True
            else:
                st.info("🎯 Combined Analysis: Daily patterns + weekly confirmation.")
                enable_daily_analysis = True
                enable_weekly_validation = True
        
        pattern_strength_min = st.slider("Pattern Strength Min:", 50, 100, 65, 5)
        
        with st.expander("🚀 Scan Settings", expanded=True):
            max_stocks = st.selectbox(
                "Stocks to Scan:",
                ["All Stocks", "First 50", "First 100", "Custom Limit"],
                index=0
            )
            if max_stocks == "Custom Limit":
                custom_limit = st.number_input("Custom Limit:", min_value=10, max_value=len(stocks_to_scan), value=50)
                stocks_limit = custom_limit
            elif max_stocks == "First 50":
                stocks_limit = 50
            elif max_stocks == "First 100":
                stocks_limit = 100
            else:
                stocks_limit = len(stocks_to_scan)
            
            show_charts = st.checkbox("Show Charts", value=True)
            show_news = st.checkbox("Show News", value=True)
            export_results = st.checkbox("Export Results", value=False)
        
        st.markdown("---")
        enhancement_options = {
            'delivery_volume': st.checkbox("📊 Delivery Volume Analysis", value=True),
            'fno_consolidation': st.checkbox("🔄 F&O Consolidation Detection", value=True),
            'breakout_pullback': st.checkbox("📈 Breakout-Pullback Patterns", value=True),
            'enhanced_sr': st.checkbox("🎯 Enhanced Support & Resistance", value=True)
        }
        
        st.markdown("---")
        scanner = ProfessionalPCSScanner()
        sentiment_data = scanner.get_market_sentiment_indicators()
        
        overall_sentiment = sentiment_data.get('overall', {})
        sentiment_level = overall_sentiment.get('sentiment', 'NEUTRAL')
        pcs_recommendation = overall_sentiment.get('pcs_recommendation', 'Moderate opportunities')
        
        sentiment_class = f"sentiment-{sentiment_level.lower()}"
        
        st.markdown(f"""
        <div class="{sentiment_class}" style="padding: 10px; border-radius: 6px; margin: 6px 0;">
            <h4 style="margin: 0 0 4px 0; color: var(--text-primary); font-size: 1rem;">
                {'🟢' if sentiment_level == 'BULLISH' else '🟡' if sentiment_level == 'NEUTRAL' else '🔴'} 
                {sentiment_level}
            </h4>
            <p style="margin: 0; font-size: 0.8rem; opacity: 0.9;">{pcs_recommendation}</p>
        </div>
        """, unsafe_allow_html=True)
        
        if 'nifty' in sentiment_data:
            nifty_data = sentiment_data['nifty']
            # delta as percentage string is fine
            st.metric("Nifty 50", f"{nifty_data['current']:.0f}", f"{nifty_data['change_1d']:+.2f}%")
        
        ist = pytz.timezone('Asia/Kolkata')
        current_time = datetime.now(ist)
        st.markdown(f"**Updated:** {current_time.strftime('%H:%M IST')}")
    
    return {
        'stocks_to_scan': stocks_to_scan[:stocks_limit],
        'rsi_min': rsi_min,
        'rsi_max': rsi_max,
        'adx_min': adx_min,
        'ma_support': ma_support,
        'ma_type': ma_type,
        'ma_tolerance': ma_tolerance,
        'min_volume_ratio': min_volume_ratio,
        'volume_breakout_ratio': volume_breakout_ratio,
        'lookback_days': lookback_days,
        'pattern_strength_min': pattern_strength_min,
        'pattern_filters': pattern_filters,
        'pattern_priority': pattern_priority,
        'analysis_mode': analysis_mode,
        'enable_daily_analysis': enable_daily_analysis,
        'enable_weekly_validation': enable_weekly_validation,
        'show_charts': show_charts,
        'show_news': show_news,
        'export_results': export_results,
        'stocks_limit': stocks_limit,
        'market_sentiment': sentiment_data,
        'enhancements': enhancement_options
    }

def create_main_scanner_tab(config):
    """Create main scanner tab with current day focus"""
    scanner = ProfessionalPCSScanner()
    st.markdown(f"**Ready to scan {len(config['stocks_to_scan'])} stocks**")
    scan_button = st.button("🚀 Start Scan", type="primary", key="main_scan", use_container_width=True)
    
    if scan_button:
        progress_bar = st.progress(0)
        status_container = st.empty()
        
        results = []
        for i, symbol in enumerate(config['stocks_to_scan']):
            progress = (i + 1) / len(config['stocks_to_scan'])
            progress_bar.progress(progress)
            
            clean_symbol = symbol.replace('.NS', '').replace('^', '')
            status_container.info(f"🔍 Analyzing {clean_symbol} ({i+1}/{len(config['stocks_to_scan'])})")
            
            try:
                data = scanner.get_stock_data(symbol, period="3mo")
                if data is None:
                    continue
                
                volume_ok, volume_ratio, volume_details = scanner.check_volume_criteria(data, config['min_volume_ratio'])
                if not volume_ok:
                    continue
                
                patterns = scanner.detect_patterns(data, symbol, config)
                if not patterns:
                    continue
                
                current_price = float(data['Close'].iloc[-1])
                current_rsi = float(data['RSI'].iloc[-1])
                current_adx = float(data['ADX'].iloc[-1])
                
                news_data = None
                if config['show_news']:
                    try:
                        news_data = scanner.get_fundamental_news(symbol, clean_symbol)
                    except Exception:
                        news_data = None
                
                enhancement_results = {}
                if config.get('enhancements', {}).get('delivery_volume', False):
                    try:
                        delivery_analysis = scanner.analyze_delivery_volume_percentage(symbol)
                        enhancement_results['delivery_volume'] = delivery_analysis
                    except Exception as e:
                        enhancement_results['delivery_volume'] = {
                            'delivery_percentage': None,
                            'delivery_analysis': f'Error: {str(e)}',
                            'delivery_signals': [],
                            'confidence': 'Low'
                        }
                
                if config.get('enhancements', {}).get('fno_consolidation', False):
                    try:
                        consolidation_analysis = scanner.detect_fno_consolidation_near_resistance(
                            data, symbol, lookback_days=20
                        )
                        enhancement_results['fno_consolidation'] = consolidation_analysis
                    except Exception as e:
                        enhancement_results['fno_consolidation'] = {
                            'consolidation_detected': False,
                            'analysis': f'Error: {str(e)}',
                            'signals': []
                        }
                
                if config.get('enhancements', {}).get('breakout_pullback', False):
                    try:
                        breakout_pullback_analysis = scanner.detect_breakout_pullback_strong_green(
                            data, lookback_days=30
                        )
                        enhancement_results['breakout_pullback'] = breakout_pullback_analysis
                    except Exception as e:
                        enhancement_results['breakout_pullback'] = {
                            'pattern_detected': False,
                            'analysis': f'Error: {str(e)}',
                            'signals': []
                        }
                
                if config.get('enhancements', {}).get('enhanced_sr', False):
                    try:
                        sr_analysis = scanner.enhanced_support_resistance_analysis(
                            data, lookback_days=50
                        )
                        enhancement_results['enhanced_sr'] = sr_analysis
                    except Exception as e:
                        enhancement_results['enhanced_sr'] = {
                            'analysis_available': False,
                            'message': f'Error: {str(e)}',
                            'support_levels': [],
                            'resistance_levels': []
                        }
                
                stock_result = {
                    'symbol': symbol,
                    'current_price': current_price,
                    'volume_ratio': volume_ratio,
                    'volume_details': volume_details,
                    'rsi': current_rsi,
                    'adx': current_adx,
                    'patterns': patterns,
                    'data': data,
                    'news_data': news_data
                }
                if enhancement_results:
                    stock_result['enhancements'] = enhancement_results
                
                results.append(stock_result)
            except Exception:
                continue
        
        progress_bar.empty()
        status_container.empty()
        
        if results:
            results.sort(key=lambda x: max(p['strength'] for p in x['patterns']), reverse=True)
            st.success(f"🎉 Found **{len(results)} stocks** with current day confirmed patterns!")
            
            total_strengths = [p['strength'] for r in results for p in r['patterns']]
            avg_strength = float(np.nanmean(total_strengths)) if total_strengths else 0.0
            current_day_breakouts = sum(1 for r in results for p in r['patterns'] if 'Current Day' in p['type'])
            high_confidence = sum(1 for r in results for p in r['patterns'] if p['confidence'] == 'HIGH')
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("🎯 Stocks Found", len(results))
            with col2:
                st.metric("🔥 Current Day", current_day_breakouts)
            with col3:
                st.metric("💪 Avg Strength", f"{avg_strength:.1f}%")
            with col4:
                st.metric("🏆 High Confidence", high_confidence)
            
            st.session_state['scan_results'] = results
    
    if st.session_state.get('scan_results'):
        results = st.session_state['scan_results']
        st.markdown("---")
        st.markdown("### 📋 Scan Results - Select Stock for Analysis")
        
        summary_data = []
        for result in results:
            max_strength = max(p['strength'] for p in result['patterns'])
            overall_confidence = 'HIGH' if max_strength >= 85 else 'MEDIUM' if max_strength >= 70 else 'LOW'
            has_current_breakout = any('Current Day' in p['type'] for p in result['patterns'])
            has_news = result.get('news_data') and result['news_data']['news_count'] > 0
            stock_key = result['symbol'].replace('.NS', '').replace('^', '')
            summary_data.append({
                'Symbol': stock_key,
                'Confidence': overall_confidence,
                'Strength': f"{max_strength:.0f}%",
                'Price': f"₹{result['current_price']:.2f}",
                'Volume': f"{result['volume_ratio']:.1f}x",
                'RSI': f"{result['rsi']:.1f}",
                'ADX': f"{result['adx']:.1f}",
                'Today': '🔥' if has_current_breakout else '',
                'News': '📰' if has_news else '',
                'Patterns': len(result['patterns'])
            })
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            st.dataframe(
                summary_df,
                use_container_width=True,
                hide_index=True,
                height=min(400, len(summary_df) * 35 + 38)
            )
            
            st.markdown("---")
            stock_options = [f"{row['Symbol']} - {row['Confidence']} ({row['Strength']})" for row in summary_data]
            if 'selected_stock_idx' not in st.session_state:
                st.session_state['selected_stock_idx'] = 0
            
            selected_display = st.selectbox(
                "🔍 Select a stock to view detailed analysis:",
                options=stock_options,
                index=st.session_state['selected_stock_idx'],
                key='stock_selector'
            )
            st.session_state['selected_stock_idx'] = stock_options.index(selected_display)
            result = results[st.session_state['selected_stock_idx']]
            
            stock_key = result['symbol'].replace('.NS', '').replace('^', '')
            max_strength = max(p['strength'] for p in result['patterns'])
            overall_confidence = 'HIGH' if max_strength >= 85 else 'MEDIUM' if max_strength >= 70 else 'LOW'
            
            st.markdown("---")
            st.markdown(f"## 📊 Detailed Analysis: **{stock_key}**")
            
            detail_tabs = st.tabs(["📈 Overview", "🎯 Patterns", "🚀 Enhancements", "📊 Charts"])
            
            with detail_tabs[0]:
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("💰 Current Price", f"₹{result['current_price']:.2f}")
                with col2:
                    volume_color = "inverse" if result['volume_ratio'] >= 2 else "normal"
                    # No delta, only value + delta_color
                    st.metric("📊 Volume Today", f"{result['volume_ratio']:.2f}x", delta_color=volume_color)
                with col3:
                    st.metric("📈 RSI", f"{result['rsi']:.1f}")
                with col4:
                    st.metric("⚡ ADX", f"{result['adx']:.1f}")
                
                current_day_data = result['data'].iloc[-1]
                trading_date = current_day_data.name.strftime('%Y-%m-%d')
                day_change = 0.0
                if current_day_data['Open'] != 0:
                    day_change = (current_day_data['Close'] - current_day_data['Open']) / current_day_data['Open'] * 100.0
                
                st.markdown(
                    f"**🗓️ Trading Date:** {trading_date} | "
                    f"**📊 Day Range:** ₹{current_day_data['Low']:.2f} - ₹{current_day_data['High']:.2f} | "
                    f"**💹 Day Change:** {day_change:+.2f}%"
                )
                
                has_news = result.get('news_data') and result['news_data']['news_count'] > 0
                if has_news:
                    news_data = result['news_data']
                    sentiment = news_data.get('overall_sentiment', 'neutral')
                    sentiment_emoji = "🟢" if sentiment == 'positive' else "🔴" if sentiment == 'negative' else "🟡"
                    st.markdown(f"### {sentiment_emoji} Today's News - {sentiment.upper()} Sentiment")
                    for news_item in news_data['news_items'][:2]:
                        relevance = news_item.get('relevance', 'low')
                        relevance_emoji = "🔥" if relevance == 'high' else "⚡" if relevance == 'medium' else "📄"
                        st.info(f"**{relevance_emoji}** {news_item['headline']}")
            
            with detail_tabs[1]:
                for pattern in result['patterns']:
                    confidence_emoji = "🟢" if pattern['confidence'] == 'HIGH' else "🟡" if pattern['confidence'] == 'MEDIUM' else "🔴"
                    st.markdown(f"### {confidence_emoji} {pattern['type']} - {pattern['confidence']} Confidence")
                    
                    pcol1, pcol2, pcol3 = st.columns(3)
                    with pcol1:
                        st.metric("Strength", f"{pattern['strength']}%")
                    with pcol2:
                        st.metric("Success Rate", f"{pattern['success_rate']}%")
                    with pcol3:
                        st.metric("PCS Fit", f"{pattern['pcs_suitability']}%")
                    
                    weekly_val = pattern.get('weekly_validation', {})
                    if isinstance(weekly_val, dict):
                        if weekly_val.get('weekly_validation', False):
                            st.success(f"📈 Weekly Confirmation: {weekly_val.get('weekly_context', '')}")
                        elif weekly_val.get('weekly_strength_bonus', 0) > 0:
                            st.warning(f"📊 Weekly Support: {weekly_val.get('weekly_context', '')}")
                    st.markdown("---")
            
            with detail_tabs[2]:
                enhancements = result.get('enhancements', {})
                
                if st.checkbox("🔍 Show Enhancement Debug Info", key=f"debug_enh_{stock_key}"):
                    st.json({
                        "enhancement_keys": list(enhancements.keys()) if enhancements else [],
                        "enhancement_data": enhancements
                    })
                
                if enhancements and len(enhancements) > 0:
                    st.markdown("### 🚀 Enhancement Analysis")
                    has_any_enhancement = False
                    
                    if 'delivery_volume' in enhancements:
                        delivery = enhancements['delivery_volume']
                        if isinstance(delivery, dict) and delivery.get('delivery_percentage') is not None:
                            has_any_enhancement = True
                            st.markdown("#### 📊 Delivery Volume Analysis")
                            ecol1, ecol2 = st.columns(2)
                            with ecol1:
                                try:
                                    delivery_pct = float(delivery.get('delivery_percentage', 0))
                                    st.metric("Estimated Delivery", f"{delivery_pct:.1f}%")
                                except (ValueError, TypeError):
                                    st.metric("Estimated Delivery", "N/A")
                            with ecol2:
                                conf = str(delivery.get('confidence', 'Low'))
                                st.metric("Confidence", conf)
                            analysis = delivery.get('delivery_analysis', 'No analysis available')
                            st.write(f"**Analysis:** {analysis}")
                            st.markdown("---")
                    
                    if 'fno_consolidation' in enhancements:
                        consolidation = enhancements['fno_consolidation']
                        if isinstance(consolidation, dict):
                            has_any_enhancement = True
                            st.markdown("#### 🔄 F&O Consolidation")
                            ecol1, ecol2 = st.columns(2)
                            with ecol1:
                                detected = consolidation.get('consolidation_detected', False)
                                st.metric("Status", "✅ Detected" if detected else "❌ Not Detected")
                            with ecol2:
                                strength = consolidation.get('consolidation_strength', 0)
                                try:
                                    strength_val = int(strength) if strength is not None else 0
                                    st.metric("Strength", f"{strength_val}/100")
                                except (ValueError, TypeError):
                                    st.metric("Strength", "N/A")
                            analysis = consolidation.get('analysis', 'No analysis available')
                            st.write(f"**Analysis:** {analysis}")
                            st.markmarkdown("---")
                    
                    if 'breakout_pullback' in enhancements:
                        breakout = enhancements['breakout_pullback']
                        if isinstance(breakout, dict):
                            has_any_enhancement = True
                            st.markdown("#### 📈 Breakout-Pullback")
                            ecol1, ecol2 = st.columns(2)
                            with ecol1:
                                detected = breakout.get('pattern_detected', False)
                                st.metric("Status", "✅ Detected" if detected else "❌ Not Detected")
                            with ecol2:
                                strength = breakout.get('pattern_strength', 0)
                                try:
                                    strength_val = int(strength) if strength is not None else 0
                                    st.metric("Strength", f"{strength_val}/100")
                                except (ValueError, TypeError):
                                    st.metric("Strength", "N/A")
                            analysis = breakout.get('analysis', 'No analysis available')
                            st.write(f"**Analysis:** {analysis}")
                            st.markdown("---")
                    
                    if 'enhanced_sr' in enhancements:
                        sr = enhancements['enhanced_sr']
                        if isinstance(sr, dict) and sr.get('analysis_available'):
                            has_any_enhancement = True
                            st.markdown("#### 🎯 Enhanced Support & Resistance")
                            ecol1, ecol2 = st.columns(2)
                            with ecol1:
                                support_count = len(sr.get('support_levels', []))
                                st.metric("Support Levels", support_count)
                            with ecol2:
                                resistance_count = len(sr.get('resistance_levels', []))
                                st.metric("Resistance Levels", resistance_count)
                            position_analysis = sr.get('position_analysis', {})
                            position_strength = position_analysis.get('position_strength', 'N/A') if isinstance(position_analysis, dict) else 'N/A'
                            st.write(f"**Position:** {position_strength}")
                            st.markdown("---")
                    
                    if not has_any_enhancement:
                        st.info("ℹ️ Enhancements enabled but no significant patterns for this stock.")
                else:
                    st.warning("⚠️ No enhancement data available. Enable enhancements in the sidebar.")
            
            with detail_tabs[3]:
                if config['show_charts']:
                    st.markdown("#### 📊 Technical Chart Analysis")
                    chart = scanner.create_tradingview_chart(
                        result['data'],
                        result['symbol'],
                        result['patterns'][0] if result['patterns'] else None
                    )
                    if chart:
                        st.plotly_chart(chart, use_container_width=True)
                else:
                    st.info("📊 Charts are disabled. Enable in sidebar settings.")
    else:
        st.warning("🔍 No current day patterns found. Try adjusting filters.")
        st.markdown("### 💡 Suggestions:")
        st.markdown("- Lower **Pattern Strength** to 50-60%")
        st.markdown("- Reduce **Volume Ratio** to 1.0x")
        st.markdown("- Expand **RSI range** to 25-85")
        st.markdown("- Check if markets traded today")

def main():
    config = create_professional_sidebar()
    tab1, tab2 = st.tabs(["🎯 Scanner", "🌍 Market Overview"])
    
    with tab1:
        create_main_scanner_tab(config)
    
    with tab2:
        st.markdown("""
        <style>
        .bloomberg-card {
            background: #0a0a0a;
            border: 1px solid #2a2a2a;
            border-left: 3px solid #ff6b00;
            padding: 16px 20px;
            margin-bottom: 12px;
            font-family: 'Courier New', monospace;
        }
        .bloomberg-card:hover {
            background: #151515;
            border-left-color: #ff8533;
        }
        .bb-symbol {
            font-size: 13px;
            color: #999;
            font-weight: 600;
            letter-spacing: 0.5px;
            text-transform: uppercase;
        }
        .bb-name {
            font-size: 11px;
            color: #666;
            margin-top: 2px;
        }
        .bb-price {
            font-size: 28px;
            font-weight: 700;
            color: #fff;
            margin: 8px 0 4px 0;
            letter-spacing: -0.5px;
        }
        .bb-change {
            font-size: 16px;
            font-weight: 600;
        }
        .bb-change.positive {
            color: #00d084;
        }
        .bb-change.negative {
            color: #ff3b69;
        }
        .bb-section {
            margin: 30px 0 15px 0;
            padding-bottom: 8px;
            border-bottom: 1px solid #2a2a2a;
        }
        .bb-section-title {
            font-size: 12px;
            color: #ff6b00;
            font-weight: 700;
            letter-spacing: 1px;
            text-transform: uppercase;
        }
        </style>
        """, unsafe_allow_html=True)
        
        scanner = ProfessionalPCSScanner()
        indices_data = scanner.get_global_indices()
        
        st.markdown('<div class="bb-section"><div class="bb-section-title">INDIAN MARKETS</div></div>', unsafe_allow_html=True)
        india_cols = st.columns(3)
        india_indices = [
            ('Nifty 50', 'NSE:NIFTY'),
            ('Bank Nifty', 'NSE:BANKNIFTY'),
            ('GIFT Nifty', 'SGX:NIFTY')
        ]
        
        for col, (index_name, symbol) in zip(india_cols, india_indices):
            with col:
                if index_name in indices_data:
                    idx = indices_data[index_name]
                    change_pct = idx['change']
                    change_class = "positive" if change_pct >= 0 else "negative"
                    change_sign = "+" if change_pct >= 0 else ""
                    st.markdown(f"""
                    <div class="bloomberg-card">
                        <div class="bb-symbol">{symbol}</div>
                        <div class="bb-name">{index_name}</div>
                        <div class="bb-price">{idx['value']:,.2f}</div>
                        <div class="bb-change {change_class}">{change_sign}{change_pct:.2f}%</div>
                    </div>
                    """, unsafe_allow_html=True)
        
        st.markdown('<div class="bb-section"><div class="bb-section-title">US MARKETS</div></div>', unsafe_allow_html=True)
        us_cols = st.columns(3)
        us_indices = [
            ('S&P 500', 'SPX INDEX'),
            ('Dow Jones', 'INDU INDEX'),
            ('Nasdaq', 'COMP INDEX')
        ]
        
        for col, (index_name, symbol) in zip(us_cols, us_indices):
            with col:
                if index_name in indices_data:
                    idx = indices_data[index_name]
                    change_pct = idx['change']
                    change_class = "positive" if change_pct >= 0 else "negative"
                    change_sign = "+" if change_pct >= 0 else ""
                    st.markdown(f"""
                    <div class="bloomberg-card">
                        <div class="bb-symbol">{symbol}</div>
                        <div class="bb-name">{index_name}</div>
                        <div class="bb-price">{idx['value']:,.0f}</div>
                        <div class="bb-change {change_class}">{change_sign}{change_pct:.2f}%</div>
                    </div>
                    """, unsafe_allow_html=True)
        
        st.markdown('<div class="bb-section"><div class="bb-section-title">ASIA MARKETS</div></div>', unsafe_allow_html=True)
        asia_cols = st.columns(3)
        asia_indices = [
            ('Nikkei 225', 'NKY INDEX'),
            ('Hang Seng', 'HSI INDEX'),
            ('Shanghai', 'SHCOMP INDEX')
        ]
        
        for col, (index_name, symbol) in zip(asia_cols, asia_indices):
            with col:
                if index_name in indices_data:
                    idx = indices_data[index_name]
                    change_pct = idx['change']
                    change_class = "positive" if change_pct >= 0 else "negative"
                    change_sign = "+" if change_pct >= 0 else ""
                    st.markdown(f"""
                    <div class="bloomberg-card">
                        <div class="bb-symbol">{symbol}</div>
                        <div class="bb-name">{index_name}</div>
                        <div class="bb-price">{idx['value']:,.0f}</div>
                        <div class="bb-change {change_class}">{change_sign}{change_pct:.2f}%</div>
                    </div>
                    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
